from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets, status, filters
from rest_framework.views import APIView
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Count, Q, Avg, Sum
from django.utils import timezone
from datetime import timedelta, datetime
import json
import io
import zipfile
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.pdfgen import canvas
from io import BytesIO
import pandas as pd
from django.http import HttpResponse
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import csv
from django.db import models

from .models import DynamicFormSchema, FormEntry, FormField, FileAttachment, FormFieldFile
from .serializers import (
    DynamicFormSchemaSerializer,
    DynamicFormSchemaCreateSerializer,
    FormEntrySerializer,
    FormEntryCreateSerializer,
    FormEntryUpdateSerializer,
    FormFieldSerializer,
    FormFieldCreateSerializer,
    FormSchemaStatisticsSerializer,
    FileAttachmentSerializer,
    FileAttachmentCreateSerializer,
    FileAttachmentUpdateSerializer,
    FormFieldFileSerializer,
    FormFieldFileCreateSerializer,
    FormFieldFileUpdateSerializer
)
from accounts.permissions import IsOrganizationAdmin
from utils.storage import S3FileManager

# Set up logging
logger = logging.getLogger(__name__)

class DynamicFormSchemaViewSet(viewsets.ModelViewSet):
    """ViewSet for DynamicFormSchema management"""
    
    queryset = DynamicFormSchema.objects.all()
    serializer_class = DynamicFormSchemaSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_active']
    search_fields = ['name', 'description']
    ordering_fields = ['name', 'created_at', 'fields_count']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return DynamicFormSchemaCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return DynamicFormSchemaUpdateSerializer
        return DynamicFormSchemaSerializer
    
    def get_permissions(self):
        """Set permissions based on action"""
        if self.action in ['create', 'destroy', 'update', 'partial_update']:
            permission_classes = [IsAuthenticated, IsOrganizationAdmin]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        user = self.request.user
        
        print(f"🔍 Schema filtering for user: {user.email} (role: {user.role}, org: {user.organization})")
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can see all schemas
            queryset = DynamicFormSchema.objects.all()
            print(f"🔍 Super admin - returning all schemas: {queryset.count()}")
        elif user.role == 'ADMIN':
            # Admin can only see schemas in their organization
            queryset = DynamicFormSchema.objects.filter(organization=user.organization)
            print(f"🔍 Admin - returning schemas for org {user.organization}: {queryset.count()}")
        else:
            # Employees can only see active schemas from their own organization
            queryset = DynamicFormSchema.objects.filter(
                organization=user.organization,
                is_active=True
            )
            print(f"🔍 Employee - returning active schemas for org {user.organization}: {queryset.count()}")
        
        return queryset
    
    def perform_create(self, serializer):
        """Set organization and created_by for new schemas"""
        user = self.request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can create schemas for any organization
            pass
        elif user.role == 'ADMIN':
            # Admin can only create schemas in their organization
            serializer.save(
                organization=user.organization,
                created_by=user
            )
            return
        
        serializer.save(created_by=user)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get form schema statistics"""
        user = request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin gets all statistics
            total_schemas = DynamicFormSchema.objects.count()
            active_schemas = DynamicFormSchema.objects.filter(is_active=True).count()
            total_entries = FormEntry.objects.count()
            completed_entries = FormEntry.objects.filter(is_completed=True).count()
            verified_entries = FormEntry.objects.filter(is_verified=True).count()
            
            # Calculate average completion time using database fields
            completed_entries_with_time = FormEntry.objects.filter(
                is_completed=True,
                tat_completion_time__isnull=False
            )
            
            # Calculate average using Python instead of database aggregation
            total_hours = 0
            count_with_time = 0
            for entry in completed_entries_with_time:
                if entry.tat_duration:
                    total_hours += entry.tat_duration
                    count_with_time += 1
            
            avg_completion_time = total_hours / count_with_time if count_with_time > 0 else 0
            
            # Recent schemas
            recent_schemas = DynamicFormSchema.objects.order_by('-created_at')[:5]
            
        else:
            # Admin/Employee gets organization-specific statistics
            org = user.organization
            total_schemas = DynamicFormSchema.objects.filter(organization=org).count()
            active_schemas = DynamicFormSchema.objects.filter(
                organization=org, 
                is_active=True
            ).count()
            total_entries = FormEntry.objects.filter(organization=org).count()
            completed_entries = FormEntry.objects.filter(
                organization=org, 
                is_completed=True
            ).count()
            verified_entries = FormEntry.objects.filter(
                organization=org, 
                is_verified=True
            ).count()
            
            # Calculate average completion time for organization
            completed_entries_with_time = FormEntry.objects.filter(
                organization=org,
                is_completed=True,
                tat_completion_time__isnull=False
            )
            
            # Calculate average using Python instead of database aggregation
            total_hours = 0
            count_with_time = 0
            for entry in completed_entries_with_time:
                if entry.tat_duration:
                    total_hours += entry.tat_duration
                    count_with_time += 1
            
            avg_completion_time = total_hours / count_with_time if count_with_time > 0 else 0
            
            # Recent schemas for organization
            recent_schemas = DynamicFormSchema.objects.filter(
                organization=org
            ).order_by('-created_at')[:5]
        
        # Serialize recent schemas separately
        recent_schemas_data = DynamicFormSchemaSerializer(recent_schemas, many=True, context={'request': request}).data
        
        data = {
            'total_schemas': total_schemas,
            'active_schemas': active_schemas,
            'total_entries': total_entries,
            'completed_entries': completed_entries,
            'verified_entries': verified_entries,
            'average_completion_time': round(avg_completion_time, 2),
            'recent_schemas': recent_schemas_data
        }
        
        serializer = FormSchemaStatisticsSerializer(data)
        return Response(serializer.data)
    
    @action(detail=True, methods=['get'])
    def entries(self, request, pk=None):
        """Get form entries for a specific schema"""
        user = request.user
        schema = self.get_object()
        
        # Check permissions
        if user.role == 'EMPLOYEE':
            # Employees can only see their own entries
            entries = FormEntry.objects.filter(
                form_schema=schema,
                employee=user
            )
        else:
            # Admin/Super admin can see all entries for the schema
            entries = FormEntry.objects.filter(form_schema=schema)
        
        serializer = FormEntrySerializer(entries, many=True, context={'request': request})
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def duplicate(self, request, pk=None):
        """Duplicate a form schema"""
        user = request.user
        original_schema = self.get_object()
        
        # Create a copy of the schema
        new_schema = DynamicFormSchema.objects.create(
            name=f"{original_schema.name} (Copy)",
            description=original_schema.description,
            organization=original_schema.organization,
            fields_definition=original_schema.fields_definition,
            max_fields=original_schema.max_fields,
            created_by=user
        )
        
        serializer = DynamicFormSchemaSerializer(new_schema, context={'request': request})
        return Response(serializer.data, status=status.HTTP_201_CREATED)

class FormEntryViewSet(viewsets.ModelViewSet):
    """ViewSet for FormEntry management"""
    
    queryset = FormEntry.objects.all()
    serializer_class = FormEntrySerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['is_completed', 'is_verified', 'form_schema', 'employee']
    search_fields = ['verification_notes', 'form_data']
    ordering_fields = ['created_at', 'case_id']
    ordering = ['-created_at']
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return FormEntryCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return FormEntryUpdateSerializer
        return FormEntrySerializer
    
    def get_permissions(self):
        """Set permissions based on action"""
        if self.action in ['destroy']:
            # Allow employees to delete their own entries, admins to delete any
            permission_classes = [IsAuthenticated]
        elif self.action in ['update', 'partial_update']:
            # Allow employees to update their own entries
            permission_classes = [IsAuthenticated]
        else:
            permission_classes = [IsAuthenticated]
        return [permission() for permission in permission_classes]

    def destroy(self, request, *args, **kwargs):
        """Custom destroy method to handle permissions"""
        instance = self.get_object()
        user = request.user
        
        # Allow employees to delete their own entries
        if user.role == 'EMPLOYEE' and instance.employee != user:
            return Response({'error': 'You can only delete your own entries'}, status=status.HTTP_403_FORBIDDEN)
        
        # Allow admins to delete any entry in their organization
        if user.role == 'ADMIN' and instance.organization != user.organization:
            return Response({'error': 'You can only delete entries in your organization'}, status=status.HTTP_403_FORBIDDEN)
        
        return super().destroy(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user
        
        # Add debugging for form data updates
        logger.info(f"=== FORM ENTRY UPDATE DEBUG ===")
        logger.info(f"📝 Updating entry: {instance.id}")
        logger.info(f"📝 Request data: {request.data}")
        logger.info(f"📝 Current form_data: {instance.form_data}")
        
        # Only allow employees to update their own entries
        if user.role == 'EMPLOYEE' and instance.employee != user:
            return Response({'error': 'You can only update your own entries'}, status=status.HTTP_403_FORBIDDEN)
        
        response = super().update(request, *args, **kwargs)
        
        # Log the updated entry
        updated_instance = self.get_object()
        logger.info(f"📝 Updated form_data: {updated_instance.form_data}")
        logger.info(f"📝 Update response: {response.data}")
        
        return response

    def partial_update(self, request, *args, **kwargs):
        instance = self.get_object()
        user = request.user
        # Only allow employees to update their own entries
        if user.role == 'EMPLOYEE' and instance.employee != user:
            return Response({'error': 'You can only update your own entries'}, status=status.HTTP_403_FORBIDDEN)
        return super().partial_update(request, *args, **kwargs)
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        user = self.request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can see all entries
            queryset = FormEntry.objects.all()
        elif user.role == 'ADMIN':
            # Admin can see all entries in their organization
            queryset = FormEntry.objects.filter(organization=user.organization)
        else:
            # Employees can see all entries in their organization (not just their own)
            queryset = FormEntry.objects.filter(organization=user.organization)
        
        # Apply custom filters
        is_out_of_tat = self.request.query_params.get('isOutOfTat')
        if is_out_of_tat and is_out_of_tat.lower() == 'true':
            # Filter for entries that are out of TAT
            queryset = queryset.filter(is_completed=False)
            # This will be further filtered by the organization's TAT limit
            # We'll need to do this in Python since it's organization-specific
        
        return queryset
    
    def list(self, request, *args, **kwargs):
        """Custom list method to handle organization-specific TAT filtering"""
        queryset = self.get_queryset()
        
        # Apply organization-specific TAT filtering
        is_out_of_tat = request.query_params.get('isOutOfTat')
        if is_out_of_tat and is_out_of_tat.lower() == 'true':
            # Filter entries that are out of TAT based on organization limit
            filtered_entries = []
            for entry in queryset:
                if entry.check_tat_status():  # This uses organization-specific TAT
                    filtered_entries.append(entry)
            queryset = queryset.filter(id__in=[entry.id for entry in filtered_entries])
        
        # Apply standard filtering and pagination
        queryset = self.filter_queryset(queryset)
        page = self.paginate_queryset(queryset)
        
        if page is not None:
            serializer = self.get_serializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)
    
    def perform_create(self, serializer):
        """Set organization and employee for new entries, and update file records"""
        try:
            user = self.request.user
            logger.info(f"🔍 Starting form entry creation for user: {user.email}")
            logger.info(f"🔍 User role: {user.role}")
            logger.info(f"🔍 User organization: {user.organization}")
            
            # Determine case_id: use provided or generate new (max+1 for org)
            data = serializer.validated_data
            org = user.organization
            case_id = data.get('case_id')
            if not case_id:
                last_case = FormEntry.objects.filter(organization=org).aggregate(max_id=models.Max('case_id'))['max_id'] or 0
                case_id = last_case + 1
            
            # Create the form entry
            entry = serializer.save(
                organization=org,
                employee=user,
                case_id=case_id
            )
            
            logger.info(f"✅ Form entry created successfully: {entry.id}")
            logger.info(f"✅ Entry ID: {entry.entry_id}")
            logger.info(f"✅ Case ID: {entry.case_id}")
            logger.info(f"✅ Form schema: {entry.form_schema}")
            
            # Update file records to point to the final form entry
            form_data = entry.form_data
            if form_data:
                logger.info(f"🔄 Updating file records for form entry: {entry.id}")
                logger.info(f"📋 Form data: {form_data}")
                
                # Find file IDs in form data
                file_ids_to_update = []
                for field_name, value in form_data.items():
                    if isinstance(value, str) and len(value) > 10 and '-' in value:
                        # This looks like a UUID - check if it's a file ID
                        try:
                            # Try to find a FormFieldFile with this ID
                            file_obj = FormFieldFile.objects.filter(id=value).first()
                            if file_obj:
                                file_ids_to_update.append((value, field_name))
                                logger.info(f"📎 Found file ID to update: {value} for field {field_name}")
                                logger.info(f"📎 Current file entry: {file_obj.form_entry}, field: {file_obj.field_name}")
                            else:
                                logger.warning(f"⚠️ File ID {value} not found in database")
                        except Exception as e:
                            logger.warning(f"⚠️ Error checking file ID {value}: {str(e)}")
                
                logger.info(f"📋 Total files to update: {len(file_ids_to_update)}")
                
                # Update file records
                updated_count = 0
                for file_id, field_name in file_ids_to_update:
                    try:
                        file_obj = FormFieldFile.objects.get(id=file_id)
                        logger.info(f"🔄 Updating file {file_id}:")
                        logger.info(f"   - Old form_entry: {file_obj.form_entry}")
                        logger.info(f"   - Old field_name: {file_obj.field_name}")
                        logger.info(f"   - New form_entry: {entry.id}")
                        logger.info(f"   - New field_name: {field_name}")
                        
                        file_obj.form_entry = entry
                        file_obj.field_name = field_name
                        file_obj.is_temporary = False  # Mark as permanent
                        file_obj.save()
                        
                        logger.info(f"✅ Updated file {file_id} to point to form entry {entry.id}")
                        updated_count += 1
                        
                        # Verify the update
                        updated_file = FormFieldFile.objects.get(id=file_id)
                        logger.info(f"✅ Verification - File {file_id} now points to entry {updated_file.form_entry}")
                        
                    except FormFieldFile.DoesNotExist:
                        logger.warning(f"⚠️ File {file_id} not found")
                    except Exception as e:
                        logger.error(f"❌ Error updating file {file_id}: {str(e)}")
                        import traceback
                        logger.error(f"❌ Traceback: {traceback.format_exc()}")
                
                logger.info(f"✅ Successfully updated {updated_count} out of {len(file_ids_to_update)} files")
            else:
                logger.info("📋 No form data to process for file updates")
            
            logger.info(f"✅ Form entry creation completed successfully: {entry.id}")
            return entry
        except Exception as e:
            logger.error(f"❌ Error in create method: {str(e)}")
            import traceback
            logger.error(f"❌ Full traceback: {traceback.format_exc()}")
            # Re-raise the exception to be handled by the view
            raise
    
    def create(self, request, *args, **kwargs):
        """Create a new form entry with enhanced error handling"""
        try:
            logger.info(f"🔍 Form entry creation request from user: {request.user.email}")
            logger.info(f"🔍 Request method: {request.method}")
            logger.info(f"🔍 Request path: {request.path}")
            logger.info(f"🔍 Request data: {request.data}")
            logger.info(f"🔍 Request content type: {request.content_type}")
            logger.info(f"🔍 Request headers: {dict(request.headers)}")
            
            serializer = self.get_serializer(data=request.data)
            
            logger.info(f"🔍 Serializer created successfully")
            
            if not serializer.is_valid():
                logger.error(f"❌ Serializer validation failed: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info(f"✅ Serializer validation passed")
            
            # Create the entry
            entry = self.perform_create(serializer)
            # Only log entry.id if entry is a model instance
            if hasattr(entry, 'id'):
                logger.info(f"✅ Form entry created successfully: {entry.id}")
            else:
                logger.info(f"✅ Form entry created successfully (no id available)")
            headers = self.get_success_headers(serializer.data)
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            logger.error(f"❌ Error in create method: {str(e)}")
            import traceback
            logger.error(f"❌ Full traceback: {traceback.format_exc()}")
            
            # Handle specific database constraint errors
            if "duplicate key value violates unique constraint" in str(e):
                if "case_id" in str(e):
                    logger.error("❌ Case ID conflict detected - this should not happen with new system")
                    logger.error("❌ This indicates a bug in the case_id generation system")
                    return Response(
                        {'error': 'Internal error with case ID generation. Please try again.'}, 
                        status=status.HTTP_500_INTERNAL_SERVER_ERROR
                    )
                elif "entry_id" in str(e):
                    logger.error("❌ Entry ID conflict detected - this indicates a race condition")
                    logger.error("❌ Multiple form entries were created simultaneously")
                    return Response(
                        {'error': 'A form entry with this ID already exists. Please try again.'}, 
                        status=status.HTTP_409_CONFLICT
                    )
                else:
                    logger.error("❌ Unknown unique constraint violation")
                    return Response(
                        {'error': 'A duplicate entry was detected. Please try again.'}, 
                        status=status.HTTP_409_CONFLICT
                    )
            
            return Response(
                {'error': 'Failed to create form entry. Please try again.'}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        """Mark a form entry as completed"""
        entry = self.get_object()
        user = request.user
        
        # Check permissions
        if user.role == 'EMPLOYEE' and entry.employee != user:
            return Response({'error': 'You can only complete your own entries'}, status=status.HTTP_403_FORBIDDEN)
        
        entry.mark_completed()
        serializer = self.get_serializer(entry)
        return Response(serializer.data)
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify a form entry"""
        entry = self.get_object()
        user = request.user
        
        # Only admins can verify entries
        if user.role == 'EMPLOYEE':
            return Response({'error': 'Only admins can verify entries'}, status=status.HTTP_403_FORBIDDEN)
        
        verification_notes = request.data.get('verification_notes', '')
        
        entry.is_verified = True
        entry.verification_notes = verification_notes
        entry.verified_by = user
        entry.verified_at = timezone.now()
        entry.save()
        
        serializer = self.get_serializer(entry)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def my_entries(self, request):
        """Get current user's form entries"""
        user = request.user
        entries = FormEntry.objects.filter(employee=user)
        
        # Apply filters from query parameters
        filters = request.query_params
        
        if filters.get('is_completed'):
            entries = entries.filter(is_completed=filters['is_completed'].lower() == 'true')
        
        if filters.get('is_verified'):
            entries = entries.filter(is_verified=filters['is_verified'].lower() == 'true')
        
        if filters.get('form_schema'):
            entries = entries.filter(form_schema_id=filters['form_schema'])
        
        serializer = self.get_serializer(entries, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def statistics(self, request):
        """Get form entry statistics"""
        user = request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin gets all statistics
            total_entries = FormEntry.objects.count()
            completed_entries = FormEntry.objects.filter(is_completed=True).count()
            verified_entries = FormEntry.objects.filter(is_verified=True).count()
            pending_entries = FormEntry.objects.filter(is_completed=False).count()
            out_of_tat_entries = FormEntry.objects.filter(
                is_completed=False,
                tat_start_time__lt=timezone.now() - timedelta(hours=24)
            ).count()
            
            # Calculate repeat cases (entries with similar data)
            repeat_cases = 0
            all_entries = FormEntry.objects.all()
            for entry in all_entries:
                similar_entries = FormEntry.objects.filter(
                    form_schema=entry.form_schema,
                    form_data__contains=entry.form_data
                ).exclude(id=entry.id)
                if similar_entries.exists():
                    repeat_cases += 1
            
            repeat_cases = repeat_cases // 2  # Divide by 2 to avoid double counting
            
        else:
            # Admin/Employee gets organization-specific statistics
            org = user.organization
            total_entries = FormEntry.objects.filter(organization=org).count()
            completed_entries = FormEntry.objects.filter(
                organization=org, 
                is_completed=True
            ).count()
            verified_entries = FormEntry.objects.filter(
                organization=org, 
                is_verified=True
            ).count()
            pending_entries = FormEntry.objects.filter(
                organization=org, 
                is_completed=False
            ).count()
            out_of_tat_entries = FormEntry.objects.filter(
                organization=org,
                is_completed=False,
                tat_start_time__lt=timezone.now() - timedelta(hours=24)
            ).count()
            
            # Calculate repeat cases for organization
            repeat_cases = 0
            org_entries = FormEntry.objects.filter(organization=org)
            for entry in org_entries:
                similar_entries = FormEntry.objects.filter(
                    organization=org,
                    form_schema=entry.form_schema,
                    form_data__contains=entry.form_data
                ).exclude(id=entry.id)
                if similar_entries.exists():
                    repeat_cases += 1
            
            repeat_cases = repeat_cases // 2  # Divide by 2 to avoid double counting
        
        # Calculate average completion time
        completed_entries_with_time = FormEntry.objects.filter(
            is_completed=True,
            tat_completion_time__isnull=False
        )
        
        total_hours = 0
        count_with_time = 0
        for entry in completed_entries_with_time:
            if entry.tat_duration:
                total_hours += entry.tat_duration
                count_with_time += 1
        
        avg_completion_time = total_hours / count_with_time if count_with_time > 0 else 0
        
        data = {
            'total_entries': total_entries,
            'completed_entries': completed_entries,
            'verified_entries': verified_entries,
            'pending_entries': pending_entries,
            'out_of_tat_entries': out_of_tat_entries,
            'repeat_cases': repeat_cases,
            'average_completion_time': round(avg_completion_time, 2)
        }
        
        return Response(data)
    
    @action(detail=False, methods=['post'])
    def advanced_filter(self, request):
        """Advanced filtering with complex queries and schema field validation"""
        try:
            print(f"🔍 Advanced Filter Debug - Request received:")
            print(f"   - Request data: {request.data}")
            print(f"   - Request user: {request.user}")
            print(f"   - Request user role: {request.user.role}")
            
            # Get filters from request data (not nested)
            filters = request.data
            user = request.user
            warnings = []
            
            # Debug: Print all received filters
            print(f"🔍 Received Filters Debug:")
            print(f"   - Request data: {request.data}")
            print(f"   - Filters: {filters}")
            print(f"   - Date range: {filters.get('date_range')}")
            print(f"   - Start date: {filters.get('start_date')}")
            print(f"   - End date: {filters.get('end_date')}")
            print(f"   - DateFrom: {filters.get('dateFrom')}")
            print(f"   - DateTo: {filters.get('dateTo')}")
            print(f"   - Page: {filters.get('page')}")
            print(f"   - Page size: {filters.get('page_size')}")
            
            # Base queryset
            if user.role == 'SUPER_ADMIN':
                queryset = FormEntry.objects.all()
            elif user.role == 'ADMIN':
                queryset = FormEntry.objects.filter(organization=user.organization)
            else:
                queryset = FormEntry.objects.filter(organization=user.organization)
            
            print(f"🔍 Base queryset count: {queryset.count()}")
            
            # Get all active schemas for the organization to validate fields
            if user.role == 'SUPER_ADMIN':
                schemas = DynamicFormSchema.objects.filter(is_active=True)
            else:
                schemas = DynamicFormSchema.objects.filter(
                    organization=user.organization,
                    is_active=True
                )
            
            print(f"🔍 Active schemas count: {schemas.count()}")
            
            # Collect all field names from schemas
            schema_fields = set()
            for schema in schemas:
                if schema.fields_definition:
                    for field in schema.fields_definition:
                        if isinstance(field, dict) and 'name' in field:
                            schema_fields.add(field['name'])
            
            print(f"🔍 Schema fields found: {schema_fields}")
            
            # Enhanced date filtering with intervals
            if filters.get('date_range'):
                now = timezone.now()
                if filters['date_range'] == 'today':
                    # Get start and end of today in user's timezone
                    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
                    today_end = today_start + timedelta(days=1)
                    queryset = queryset.filter(
                        created_at__gte=today_start,
                        created_at__lt=today_end
                    )
                    print(f"🔍 Today Filter Debug:")
                    print(f"   - Today start: {today_start}")
                    print(f"   - Today end: {today_end}")
                    print(f"   - Current time: {now}")
                    print(f"   - Query count after today filter: {queryset.count()}")
                elif filters['date_range'] == 'week':
                    week_ago = now - timedelta(days=7)
                    queryset = queryset.filter(created_at__gte=week_ago)
                elif filters['date_range'] == 'month':
                    month_ago = now - timedelta(days=30)
                    queryset = queryset.filter(created_at__gte=month_ago)
                elif filters['date_range'] == 'quarter':
                    quarter_ago = now - timedelta(days=90)
                    queryset = queryset.filter(created_at__gte=quarter_ago)
                elif filters['date_range'] == 'year':
                    year_ago = now - timedelta(days=365)
                    queryset = queryset.filter(created_at__gte=year_ago)
            
            # Custom date range
            if filters.get('start_date'):
                queryset = queryset.filter(created_at__gte=filters['start_date'])
            
            if filters.get('end_date'):
                queryset = queryset.filter(created_at__lte=filters['end_date'])
            
            # Also check for dateFrom/dateTo (frontend compatibility)
            if filters.get('dateFrom'):
                queryset = queryset.filter(created_at__gte=filters['dateFrom'])
            
            if filters.get('dateTo'):
                queryset = queryset.filter(created_at__lte=filters['dateTo'])
            
            # Month and Year filtering
            if filters.get('month'):
                queryset = queryset.filter(created_at__month=filters['month'])
            
            if filters.get('year'):
                queryset = queryset.filter(created_at__year=filters['year'])
            
            # Employee filtering
            if filters.get('employee_name'):
                queryset = queryset.filter(
                    Q(employee__first_name__icontains=filters['employee_name']) |
                    Q(employee__last_name__icontains=filters['employee_name'])
                )
            
            # Schema/Case type filtering
            if filters.get('case_type'):
                queryset = queryset.filter(form_schema__name__icontains=filters['case_type'])
            
            # Status filtering
            if filters.get('status'):
                status_value = filters['status'].lower()
                if status_value == 'completed':
                    queryset = queryset.filter(is_completed=True)
                elif status_value == 'pending':
                    queryset = queryset.filter(is_completed=False, is_verified=False)
                elif status_value == 'verified':
                    queryset = queryset.filter(is_verified=True)
            
            # TAT filtering (Out of TAT)
            if filters.get('is_out_of_tat') is not None:
                is_out_of_tat = filters['is_out_of_tat']
                # Handle both string and boolean values
                if isinstance(is_out_of_tat, str):
                    is_out_of_tat = is_out_of_tat.lower() == 'true'
                else:
                    is_out_of_tat = bool(is_out_of_tat)
                
                # Get current time for TAT calculation
                now = timezone.now()
                
                print(f"🔍 TAT Filter Debug:")
                print(f"   - Original value: {filters['is_out_of_tat']}")
                print(f"   - Processed value: {is_out_of_tat}")
                print(f"   - Current time: {now}")
                
                if is_out_of_tat:
                    # Filter for entries that are out of TAT using schema-specific limits
                    # We'll filter in Python since we need schema-specific TAT limits
                    out_of_tat_entries = []
                    for entry in queryset:
                        if entry.check_tat_status():
                            out_of_tat_entries.append(entry.id)
                    queryset = queryset.filter(id__in=out_of_tat_entries)
                    print(f"   - Filtering for OUT OF TAT entries using schema-specific limits")
                else:
                    # Filter for entries that are within TAT using schema-specific limits
                    within_tat_entries = []
                    for entry in queryset:
                        if not entry.check_tat_status():
                            within_tat_entries.append(entry.id)
                    queryset = queryset.filter(id__in=within_tat_entries)
                    print(f"   - Filtering for WITHIN TAT entries using schema-specific limits")
                
                print(f"   - Query count after TAT filter: {queryset.count()}")
            
            # Check for business-specific filters that might not exist in schema
            business_filters = [
                'bank_nbfc_name', 'location', 'product_type', 'case_status',
                'field_verifier_name', 'back_office_executive_name', 'is_repeat_case'
            ]
            
            warnings = []
            for filter_name in business_filters:
                if filters.get(filter_name) and not any(field.lower() == filter_name.lower() for field in schema_fields):
                    field_display_name = filter_name.replace('_', ' ').title()
                    warnings.append(f"Filter field '{field_display_name}' is not present in your organization's form schemas")
            
            # Apply business-specific filters if they exist in schema (case-insensitive)
            if filters.get('bank_nbfc_name') and any(field.lower() == 'bank_nbfc_name' for field in schema_fields):
                queryset = queryset.filter(form_data__bank_nbfc_name__icontains=filters['bank_nbfc_name'])
            
            if filters.get('location') and any(field.lower() == 'location' for field in schema_fields):
                queryset = queryset.filter(form_data__location__icontains=filters['location'])
            
            if filters.get('product_type') and any(field.lower() == 'product_type' for field in schema_fields):
                queryset = queryset.filter(form_data__product_type__icontains=filters['product_type'])
            
            if filters.get('case_status') and any(field.lower() == 'case_status' for field in schema_fields):
                queryset = queryset.filter(form_data__case_status__icontains=filters['case_status'])
            
            if filters.get('field_verifier_name') and any(field.lower() == 'field_verifier_name' for field in schema_fields):
                queryset = queryset.filter(form_data__field_verifier_name__icontains=filters['field_verifier_name'])
            
            if filters.get('back_office_executive_name') and any(field.lower() == 'back_office_executive_name' for field in schema_fields):
                queryset = queryset.filter(form_data__back_office_executive_name__icontains=filters['back_office_executive_name'])
            
            if filters.get('is_repeat_case') and any(field.lower() == 'is_repeat_case' for field in schema_fields):
                is_repeat = filters['is_repeat_case'].lower() == 'true'
                queryset = queryset.filter(form_data__is_repeat_case=is_repeat)
            
            # Search in all fields
            if filters.get('search'):
                search_term = filters['search']
                print(f"🔍 Search Debug - Term: {search_term}")
                
                # Create a comprehensive search query
                search_query = Q()
                
                # Search in basic fields
                search_query |= Q(employee__first_name__icontains=search_term)
                search_query |= Q(employee__last_name__icontains=search_term)
                search_query |= Q(employee__email__icontains=search_term)
                search_query |= Q(form_schema__name__icontains=search_term)
                search_query |= Q(verification_notes__icontains=search_term)
                search_query |= Q(case_id__icontains=search_term)
                
                # Search in form_data JSON field (all dynamic fields)
                search_query |= Q(form_data__icontains=search_term)
                
                # Search in filtered_form_data if available
                # Note: This might not work directly since it's a computed field
                # We'll handle this by searching the original form_data
                
                # Apply the search query
                queryset = queryset.filter(search_query)
                
                print(f"🔍 Search Debug - Query count after search: {queryset.count()}")
                
                # Additional search in specific form fields if needed
                # This searches for the term in any JSON key or value
                if search_term:
                    # Search in JSON fields more comprehensively
                    # This will search in both keys and values of the JSON
                    json_search_query = Q()
                    
                    # Search in form_data JSON (both keys and values)
                    json_search_query |= Q(form_data__icontains=search_term)
                    
                    # Apply JSON search
                    queryset = queryset.filter(json_search_query)
                    
                    print(f"🔍 Search Debug - Final count after JSON search: {queryset.count()}")
            
            # Ensure proper ordering by case_id (ascending) for consistent display
            # First, handle any entries with null case_ids by assigning them
            null_case_entries = queryset.filter(case_id__isnull=True)
            for entry in null_case_entries:
                # Force save to trigger case_id generation
                entry.save()
            
            # Order by case_id ascending for proper numerical order
            queryset = queryset.order_by('case_id')
            
            # Pagination
            # Extract pagination parameters from request data
            page = filters.get('page', 1)
            page_size = filters.get('page_size', 20)
            
            print(f"🔍 Pagination Debug:")
            print(f"   - Page: {page}")
            print(f"   - Page size: {page_size}")
            print(f"   - Final queryset count: {queryset.count()}")
            
            # Convert to integers
            try:
                page = int(page)
                page_size = int(page_size)
            except (ValueError, TypeError):
                page = 1
                page_size = 20
            
            # Calculate offset for manual pagination
            offset = (page - 1) * page_size
            
            # Apply pagination manually
            paginated_queryset = queryset[offset:offset + page_size]
            total_count = queryset.count()
            
            try:
                # Use select_related to optimize database queries
                paginated_queryset = paginated_queryset.select_related(
                    'employee', 'organization', 'form_schema', 'verified_by'
                ).prefetch_related('field_files')
                
                serializer = self.get_serializer(paginated_queryset, many=True)
                response_data = {
                    'count': total_count,
                    'next': f"?page={page + 1}" if offset + page_size < total_count else None,
                    'previous': f"?page={page - 1}" if page > 1 else None,
                    'results': serializer.data
                }
                
                if warnings:
                    response_data['warnings'] = warnings
                
                print(f"🔍 Success - Returning paginated response with {len(serializer.data)} items")
                print(f"🔍 Total count: {total_count}, Page: {page}, Page size: {page_size}")
                return Response(response_data)
                
            except Exception as e:
                print(f"❌ Error during serialization/pagination: {str(e)}")
                import traceback
                print(f"❌ Traceback: {traceback.format_exc()}")
                return Response(
                    {'error': f'Serialization error: {str(e)}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
        except Exception as e:
            print(f"❌ Advanced Filter Error: {str(e)}")
            import traceback
            print(f"❌ Traceback: {traceback.format_exc()}")
            return Response(
                {'error': f'Filter error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def counts(self, request):
        """Get case counts for different time periods with optional filtering"""
        try:
            user = request.user
            now = timezone.now()
            
            # Base queryset
            if user.role == 'SUPER_ADMIN':
                queryset = FormEntry.objects.all()
            elif user.role == 'ADMIN':
                queryset = FormEntry.objects.filter(organization=user.organization)
            else:
                queryset = FormEntry.objects.filter(organization=user.organization)
            
            # Apply filters from query parameters if provided
            filters = request.query_params
            
            if filters.get('search'):
                search_term = filters['search']
                queryset = queryset.filter(
                    Q(form_data__icontains=search_term) |
                    Q(employee__first_name__icontains=search_term) |
                    Q(employee__last_name__icontains=search_term) |
                    Q(form_schema__name__icontains=search_term)
                )
            
            if filters.get('status'):
                status_value = filters['status'].lower()
                if status_value == 'completed':
                    queryset = queryset.filter(is_completed=True)
                elif status_value == 'pending':
                    queryset = queryset.filter(is_completed=False)
                elif status_value == 'verified':
                    queryset = queryset.filter(is_verified=True)
            
            if filters.get('dateFrom'):
                queryset = queryset.filter(created_at__gte=filters['dateFrom'])
            
            if filters.get('dateTo'):
                queryset = queryset.filter(created_at__lte=filters['dateTo'])
            
            if filters.get('employee'):
                queryset = queryset.filter(
                    Q(employee__first_name__icontains=filters['employee']) |
                    Q(employee__last_name__icontains=filters['employee'])
                )
            
            if filters.get('caseType'):
                queryset = queryset.filter(form_schema__name__icontains=filters['caseType'])
            
            # Calculate time periods
            week_ago = now - timedelta(days=7)
            month_ago = now - timedelta(days=30)
            year_ago = now - timedelta(days=365)
            
            # Get counts
            total = queryset.count()
            this_week = queryset.filter(created_at__gte=week_ago).count()
            this_month = queryset.filter(created_at__gte=month_ago).count()
            this_year = queryset.filter(created_at__gte=year_ago).count()
            
            return Response({
                'total': total,
                'thisWeek': this_week,
                'thisMonth': this_month,
                'thisYear': this_year
            })
            
        except Exception as e:
            return Response(
                {'error': f'Count error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=False, methods=['get'])
    def date_range_count(self, request):
        """Get case count for a specific date range"""
        try:
            user = request.user
            start_date = request.query_params.get('start_date')
            end_date = request.query_params.get('end_date')
            
            if not start_date or not end_date:
                return Response(
                    {'error': 'Both start_date and end_date are required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Base queryset
            if user.role == 'SUPER_ADMIN':
                queryset = FormEntry.objects.all()
            elif user.role == 'ADMIN':
                queryset = FormEntry.objects.filter(organization=user.organization)
            else:
                queryset = FormEntry.objects.filter(organization=user.organization)
            
            # Apply date range filter
            queryset = queryset.filter(
                created_at__gte=start_date,
                created_at__lte=end_date
            )
            
            count = queryset.count()
            
            return Response({
                'count': count,
                'start_date': start_date,
                'end_date': end_date,
                'date_range': f"{start_date} to {end_date}"
            })
            
        except Exception as e:
            return Response(
                {'error': f'Date range count error: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )

    @action(detail=True, methods=['get'])
    def download(self, request, pk=None):
        """Download form entry as PDF"""
        entry = self.get_object()
        user = request.user
        
        # Check permissions
        if user.role == 'EMPLOYEE' and entry.employee != user:
            return Response({'error': 'You can only download your own entries'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            # Generate PDF content (simplified for now)
            
            buffer = BytesIO()
            p = canvas.Canvas(buffer, pagesize=letter)
            
            # Add content to PDF
            p.drawString(100, 750, f"Form Entry: {entry.id}")
            p.drawString(100, 730, f"Employee: {entry.employee.get_full_name()}")
            p.drawString(100, 710, f"Created: {entry.created_at}")
            p.drawString(100, 690, f"Status: {entry.get_status_display()}")
            
            # Add form data
            y_position = 650
            for key, value in entry.form_data.items():
                p.drawString(100, y_position, f"{key}: {value}")
                y_position -= 20
                if y_position < 50:
                    p.showPage()
                    y_position = 750
            
            p.save()
            buffer.seek(0)
            
            response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="entry_{entry.id}.pdf"'
            return response
            
        except Exception as e:
            return Response({'error': f'Failed to generate PDF: {str(e)}'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=True, methods=['get'])
    def view_details(self, request, pk=None):
        """Get detailed view of form entry"""
        entry = self.get_object()
        user = request.user
        
        # Check permissions
        if user.role == 'EMPLOYEE' and entry.employee != user:
            return Response({'error': 'You can only view your own entries'}, status=status.HTTP_403_FORBIDDEN)
        
        # Get form schema details
        schema = entry.form_schema
        schema_data = {
            'id': schema.id,
            'name': schema.name,
            'description': schema.description,
            'fields_definition': schema.fields_definition
        }
        
        # Get file attachments
        attachments = FileAttachment.objects.filter(form_entry=entry)
        attachment_data = FileAttachmentSerializer(attachments, many=True).data
        
        response_data = {
            'entry': self.get_serializer(entry).data,
            'schema': schema_data,
            'attachments': attachment_data
        }
        
        return Response(response_data)
    
    @action(detail=True, methods=['put'])
    def update_status(self, request, pk=None):
        """Update entry status (complete/verify)"""
        entry = self.get_object()
        user = request.user
        new_status = request.data.get('status')
        
        if not new_status:
            return Response({'error': 'Status is required'}, status=status.HTTP_400_BAD_REQUEST)
        
        # Check permissions
        if user.role == 'EMPLOYEE' and entry.employee != user:
            return Response({'error': 'You can only update your own entries'}, status=status.HTTP_403_FORBIDDEN)
        
        # Update status based on action
        if new_status == 'completed':
            if user.role == 'EMPLOYEE':
                # Employees can mark their own entries as completed
                entry.is_completed = True
                entry.tat_completion_time = timezone.now()
                entry.save()
            else:
                return Response({'error': 'Only employees can mark entries as completed'}, status=status.HTTP_403_FORBIDDEN)
        
        elif new_status == 'verified':
            if user.role in ['ADMIN', 'SUPER_ADMIN']:
                # Only admins can verify entries
                entry.is_verified = True
                entry.verified_by = user
                entry.verified_at = timezone.now()
                entry.verification_notes = request.data.get('verification_notes', '')
                entry.save()
            else:
                return Response({'error': 'Only admins can verify entries'}, status=status.HTTP_403_FORBIDDEN)
        
        elif new_status == 'pending':
            # Reset to pending (only for admins)
            if user.role in ['ADMIN', 'SUPER_ADMIN']:
                entry.is_completed = False
                entry.is_verified = False
                entry.verified_by = None
                entry.verified_at = None
                entry.verification_notes = ''
                entry.save()
            else:
                return Response({'error': 'Only admins can reset entries to pending'}, status=status.HTTP_403_FORBIDDEN)
        
        else:
            return Response({'error': 'Invalid status'}, status=status.HTTP_400_BAD_REQUEST)
        
        serializer = self.get_serializer(entry)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def test(self, request):
        """Test endpoint to verify the viewset is working"""
        logger.info(f"🔍 Test endpoint called by user: {request.user.email}")
        return Response({'message': 'Test endpoint working', 'user': request.user.email})
    
    @action(detail=False, methods=['post'])
    def test_create(self, request):
        """Test create endpoint with minimal data"""
        logger.info(f"🔍 Test create called by user: {request.user.email}")
        logger.info(f"🔍 Test create data: {request.data}")
        
        try:
            # Try to create with minimal data
            test_data = {
                'form_schema': request.data.get('form_schema'),
                'form_data': request.data.get('form_data', {})
            }
            
            serializer = self.get_serializer(data=test_data)
            if serializer.is_valid():
                entry = self.perform_create(serializer)
                return Response({'message': 'Test create successful', 'entry_id': entry.id})
            else:
                return Response({'message': 'Test create failed', 'errors': serializer.errors})
                
        except Exception as e:
            logger.error(f"❌ Test create error: {str(e)}")
            return Response({'message': 'Test create error', 'error': str(e)})

class FormFieldViewSet(viewsets.ModelViewSet):
    """ViewSet for FormField management"""
    
    queryset = FormField.objects.all()
    serializer_class = FormFieldSerializer
    permission_classes = [IsAuthenticated, IsOrganizationAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['field_type', 'is_active', 'is_required']
    search_fields = ['name', 'display_name']
    ordering_fields = ['name', 'order', 'created_at']
    ordering = ['order', 'name']
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return FormFieldCreateSerializer
        return FormFieldSerializer
    
    def get_queryset(self):
        """Filter queryset based on user organization"""
        user = self.request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can see all fields
            return FormField.objects.all()
        else:
            # Admin can only see fields in their organization
            return FormField.objects.filter(organization=user.organization)
    
    def perform_create(self, serializer):
        """Set organization and created_by for new fields"""
        user = self.request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can create fields for any organization
            pass
        else:
            # Admin can only create fields in their organization
            serializer.save(
                organization=user.organization,
                created_by=user
            )
            return
        
        serializer.save(created_by=user)

class FileAttachmentViewSet(viewsets.ModelViewSet):
    """ViewSet for file attachments"""
    
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['form_entry', 'is_verified', 'file_type', 'uploaded_by']
    search_fields = ['original_filename', 'description', 'verification_notes']
    ordering_fields = ['uploaded_at', 'file_size', 'original_filename']
    ordering = ['-uploaded_at']
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        user = self.request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can see all files
            return FileAttachment.objects.all()
        elif user.role == 'ADMIN':
            # Admin can see files in their organization
            return FileAttachment.objects.filter(form_entry__organization=user.organization)
        else:
            # Employees can only see their own files
            return FileAttachment.objects.filter(uploaded_by=user)
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return FileAttachmentCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return FileAttachmentUpdateSerializer
        return FileAttachmentSerializer
    
    def get_serializer_context(self):
        """Add request to serializer context"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def perform_create(self, serializer):
        """Set uploaded_by to current user"""
        serializer.save(uploaded_by=self.request.user)
    
    def create(self, request, *args, **kwargs):
        """Create a new file attachment with enhanced logging"""
        logger.info(f"File upload request from user {request.user.email}")
        logger.info(f"Request data keys: {list(request.data.keys())}")
        logger.info(f"Request files keys: {list(request.FILES.keys())}")
        
        try:
            serializer = self.get_serializer(data=request.data)
            serializer.is_valid(raise_exception=True)
            
            # Log file information
            file_obj = request.FILES.get('file')
            if file_obj:
                logger.info(f"File details - Name: {file_obj.name}, Size: {file_obj.size}, Type: {file_obj.content_type}")
            else:
                logger.warning("No file found in request")
            
            # Test S3 connection before upload
            s3_connection_ok = S3FileManager.test_s3_connection()
            logger.info(f"S3 connection test: {'OK' if s3_connection_ok else 'FAILED'}")
            
            self.perform_create(serializer)
            headers = self.get_success_headers(serializer.data)
            
            logger.info(f"File attachment created successfully: {serializer.data.get('id')}")
            return Response(serializer.data, status=status.HTTP_201_CREATED, headers=headers)
            
        except Exception as e:
            logger.error(f"Error creating file attachment: {str(e)}")
            logger.error(f"Request user: {request.user.email}")
            logger.error(f"Request data: {request.data}")
            raise

class FormFieldFileViewSet(viewsets.ModelViewSet):
    """ViewSet for form field file uploads"""
    
    queryset = FormFieldFile.objects.all()
    serializer_class = FormFieldFileSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['form_entry', 'field_name', 'is_verified', 'uploaded_by']
    search_fields = ['original_filename', 'description', 'verification_notes']
    ordering_fields = ['uploaded_at', 'file_size', 'original_filename']
    ordering = ['-uploaded_at']
    
    def retrieve(self, request, *args, **kwargs):
        """Get a specific form field file with debugging"""
        try:
            file_id = kwargs.get('pk')
            print(f"🔍 File request for ID: {file_id}")
            print(f"🔍 Request user: {request.user.email}")
            print(f"🔍 Request path: {request.path}")
            
            file_obj = self.get_object()
            print(f"🔍 File found: {file_obj.id}")
            print(f"🔍 File name: {file_obj.original_filename}")
            print(f"🔍 File URL: {file_obj.file.url if hasattr(file_obj, 'file') and file_obj.file else 'No file URL'}")
            
            serializer = self.get_serializer(file_obj)
            return Response(serializer.data)
        except Exception as e:
            print(f"❌ File retrieval error: {str(e)}")
            print(f"❌ Error type: {type(e)}")
            return Response(
                {'error': f'File not found: {str(e)}'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    def get_serializer_class(self):
        """Return appropriate serializer based on action"""
        if self.action == 'create':
            return FormFieldFileCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return FormFieldFileUpdateSerializer
        return FormFieldFileSerializer
    
    def get_serializer_context(self):
        """Add request to serializer context"""
        context = super().get_serializer_context()
        context['request'] = self.request
        return context
    
    def get_queryset(self):
        """Filter queryset based on user role"""
        user = self.request.user
        
        if user.role == 'SUPER_ADMIN':
            # Super admin can see all files
            return FormFieldFile.objects.all()
        elif user.role == 'ADMIN':
            # Admin can see files in their organization
            return FormFieldFile.objects.filter(form_entry__organization=user.organization)
        else:
            # Employees can only see their own files
            return FormFieldFile.objects.filter(uploaded_by=user)
    
    def perform_create(self, serializer):
        """Set uploaded_by for new files"""
        user = self.request.user
        serializer.save(uploaded_by=user)
    
    def create(self, request, *args, **kwargs):
        """Create a new form field file with detailed logging"""
        import logging
        logger = logging.getLogger(__name__)
        
        logger.info("=== FORM FIELD FILE UPLOAD DEBUG ===")
        logger.info(f"📥 Received file upload request")
        logger.info(f"📥 Request method: {request.method}")
        logger.info(f"📥 Request user: {request.user}")
        logger.info(f"📥 Request data keys: {list(request.data.keys())}")
        logger.info(f"📥 Request files: {list(request.FILES.keys()) if request.FILES else 'No files'}")
        
        # Log form data
        for key, value in request.data.items():
            if key != 'file':  # Don't log the actual file content
                logger.info(f"📥 Form data - {key}: {value}")
        
        # Log file information
        if 'file' in request.FILES:
            file_obj = request.FILES['file']
            logger.info(f"📁 File details:")
            logger.info(f"📁   Name: {file_obj.name}")
            logger.info(f"📁   Size: {file_obj.size} bytes")
            logger.info(f"📁   Type: {file_obj.content_type}")
            logger.info(f"📁   Content-Type: {file_obj.content_type}")
        else:
            logger.warning("⚠️ No file found in request.FILES")
        
        try:
            # Validate the data
            logger.info("🔍 Validating request data...")
            serializer = self.get_serializer(data=request.data)
            if not serializer.is_valid():
                logger.error("❌ Validation failed!")
                logger.error(f"❌ Validation errors: {serializer.errors}")
                return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
            
            logger.info("✅ Validation successful")
            logger.info(f"✅ Validated data: {serializer.validated_data}")
            
            # Get the validated data
            validated_data = serializer.validated_data
            logger.info(f"📋 Form entry ID: {validated_data.get('form_entry')}")
            logger.info(f"📋 Field name: {validated_data.get('field_name')}")
            logger.info(f"📋 Original filename: {validated_data.get('original_filename')}")
            logger.info(f"📋 File type: {validated_data.get('file_type')}")
            logger.info(f"📋 File size: {validated_data.get('file_size')}")
            
            # Create the file object using perform_create
            logger.info("💾 Creating file object...")
            file_obj = serializer.save()
            logger.info(f"✅ File object created with ID: {file_obj.id}")
            
            # Log the file URL and S3 details
            if hasattr(file_obj, 'file') and file_obj.file:
                logger.info(f"✅ File URL: {file_obj.file.url}")
                logger.info(f"✅ File name: {file_obj.file.name}")
                
                # Log S3 bucket details
                try:
                    if hasattr(file_obj.file, 'storage') and hasattr(file_obj.file.storage, 'bucket_name'):
                        logger.info(f"🌐 S3 Bucket: {file_obj.file.storage.bucket_name}")
                        logger.info(f"🌐 S3 Key: {file_obj.file.name}")
                        
                        # Try to get the full S3 URL
                        if hasattr(file_obj.file.storage, 'url'):
                            s3_url = file_obj.file.storage.url(file_obj.file.name)
                            logger.info(f"🌐 Full S3 URL: {s3_url}")
                        else:
                            logger.info(f"🌐 S3 URL method not available")
                    else:
                        logger.info(f"🌐 Storage details not available")
                except Exception as e:
                    logger.error(f"❌ Error getting S3 details: {str(e)}")
            
            # Store the S3 URL for easier access
            if hasattr(file_obj, 'file') and file_obj.file:
                try:
                    s3_url = file_obj.file.url
                    file_obj.s3_url = s3_url
                    file_obj.save()
                    logger.info(f"💾 Stored S3 URL: {s3_url}")
                except Exception as e:
                    logger.warning(f"⚠️ Could not store S3 URL: {str(e)}")
            
            # Return the response
            response_serializer = FormFieldFileSerializer(file_obj, context={'request': request})
            response_data = response_serializer.data
            logger.info(f"✅ Upload successful! File ID: {file_obj.id}")
            logger.info(f"✅ Response data: {response_data}")
            logger.info(f"✅ File URL: {file_obj.file.url if hasattr(file_obj, 'file') and file_obj.file else 'No file URL'}")
            logger.info(f"✅ File name: {file_obj.original_filename}")
            logger.info(f"✅ File size: {file_obj.file_size}")
            
            return Response(response_data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"❌ Error during file upload: {str(e)}")
            logger.error(f"❌ Error type: {type(e)}")
            import traceback
            logger.error(f"❌ Traceback: {traceback.format_exc()}")
            raise
    
    @action(detail=True, methods=['post'])
    def verify(self, request, pk=None):
        """Verify a form field file"""
        file_obj = self.get_object()
        user = request.user
        
        # Check permissions
        if user.role == 'EMPLOYEE':
            return Response(
                {'error': 'Employees cannot verify files'}, 
                status=status.HTTP_403_FORBIDDEN
            )
        
        verification_notes = request.data.get('verification_notes', '')
        
        file_obj.is_verified = True
        file_obj.verification_notes = verification_notes
        file_obj.verified_by = user
        file_obj.verified_at = timezone.now()
        file_obj.save()
        
        serializer = self.get_serializer(file_obj)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def by_form_entry(self, request):
        """Get files for a specific form entry"""
        form_entry_id = request.query_params.get('form_entry_id')
        if not form_entry_id:
            return Response(
                {'error': 'form_entry_id parameter is required'}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        files = FormFieldFile.objects.filter(form_entry_id=form_entry_id)
        serializer = self.get_serializer(files, many=True)
        return Response(serializer.data)

class FormEntryUploadView(APIView):
    """Handle file uploads for form entries with enhanced logging"""
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Upload a file for a form entry"""
        logger.info(f"Form entry upload request from user {request.user.email}")
        
        try:
            file_obj = request.FILES.get('file')
            employee_id = request.data.get('employee')
            
            logger.info(f"File upload details - Employee: {employee_id}")
            logger.info(f"Request data: {request.data}")
            logger.info(f"Request files: {list(request.FILES.keys())}")
            
            if not file_obj:
                logger.warning("No file provided in upload request")
                return Response(
                    {'error': 'No file provided'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            logger.info(f"File details - Name: {file_obj.name}, Size: {file_obj.size}, Type: {file_obj.content_type}")
            
            # Validate file type
            allowed_types = [
                'image/jpeg', 'image/png', 'image/gif', 
                'application/pdf', 'application/msword',
                'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
            ]
            
            if file_obj.content_type not in allowed_types:
                logger.warning(f"Invalid file type: {file_obj.content_type}")
                return Response(
                    {'error': 'Invalid file type. Allowed: JPEG, PNG, GIF, PDF, DOC, DOCX'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Validate file size (5MB limit)
            if file_obj.size > 5 * 1024 * 1024:
                logger.warning(f"File too large: {file_obj.size} bytes")
                return Response(
                    {'error': 'File size must be less than 5MB'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            # Test S3 connection
            s3_connection_ok = S3FileManager.test_s3_connection()
            logger.info(f"S3 connection test: {'OK' if s3_connection_ok else 'FAILED'}")
            
            # Create file attachment
            attachment = FileAttachment.objects.create(
                original_filename=file_obj.name,
                file=file_obj,
                file_size=file_obj.size,
                file_type=file_obj.content_type,
                uploaded_by=request.user,
                description=f"Uploaded by {request.user.get_full_name() or request.user.email}"
            )
            
            logger.info(f"File attachment created: {attachment.id}")
            logger.info(f"File saved to: {attachment.file.name}")
            
            # Test if file was actually saved
            if hasattr(attachment.file, 'storage'):
                file_exists = attachment.file.storage.exists(attachment.file.name)
                logger.info(f"File exists in storage: {file_exists}")
            
            serializer = FileAttachmentSerializer(attachment, context={'request': request})
            logger.info(f"File upload completed successfully: {attachment.id}")
            
            return Response(serializer.data, status=status.HTTP_201_CREATED)
            
        except Exception as e:
            logger.error(f"Error in file upload: {str(e)}")
            logger.error(f"Request user: {request.user.email}")
            logger.error(f"Request data: {request.data}")
            return Response(
                {'error': f'Upload failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

class FormEntryExportView(APIView):
    """
    Advanced export functionality for form entries
    Supports Excel with file references, PDF with attachments, and CSV
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Export form entries based on filters"""
        user = request.user
        export_format = request.data.get('format', 'excel')
        filters = request.data.get('filters', {})
        options = request.data.get('options', {})
        
        logger.info(f"Export request from user {user.email}: format={export_format}, filters={filters}")
        
        # Get filtered entries
        if user.role == 'SUPER_ADMIN':
            organization = None
        else:
            organization = user.organization
        
        entries = self.get_filtered_entries(filters, organization)
        logger.info(f"Found {entries.count()} entries for export")
        
        # Generate filename with date range information
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        date_range_text = self.get_date_range_text(filters)
        file_name = f"form_entries_{date_range_text}_{timestamp}"
        
        logger.info(f"Exporting {entries.count()} entries with date range: {date_range_text}")
        
        if export_format == 'excel':
            return self.export_to_excel(entries, options, file_name)
        elif export_format == 'pdf':
            return self.export_to_pdf(entries, options, file_name)
        elif export_format == 'csv':
            return self.export_to_csv(entries, options, file_name)
        else:
            return Response({'error': 'Unsupported export format'}, status=status.HTTP_400_BAD_REQUEST)
    
    def get_date_range_text(self, filters):
        """Generate date range text for filename"""
        date_range = filters.get('date_range', 'all')
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        month = filters.get('month')
        year = filters.get('year')
        
        if date_range == 'today':
            return 'today'
        elif date_range == 'week':
            return 'last_7_days'
        elif date_range == 'month':
            return 'last_30_days'
        elif date_range == 'quarter':
            return 'last_90_days'
        elif date_range == 'year':
            return 'last_365_days'
        elif date_range == 'custom' and start_date and end_date:
            return f"{start_date}_to_{end_date}"
        elif start_date and end_date:
            return f"{start_date}_to_{end_date}"
        elif start_date:
            return f"from_{start_date}"
        elif end_date:
            return f"until_{end_date}"
        elif month and year:
            return f"{year}_month_{month}"
        elif year:
            return f"year_{year}"
        else:
            return 'all_entries'
    
    def get_filtered_entries(self, filters, organization):
        """Get filtered entries with date range support"""
        queryset = FormEntry.objects.all()
        
        # Filter by organization
        if organization:
            queryset = queryset.filter(organization=organization)
        
        # Date range filtering
        date_range = filters.get('date_range', 'all')
        custom_start_date = filters.get('custom_start_date')
        custom_end_date = filters.get('custom_end_date')
        
        if date_range == 'last_7_days':
            start_date = timezone.now() - timedelta(days=7)
            queryset = queryset.filter(created_at__gte=start_date)
        elif date_range == 'last_30_days':
            start_date = timezone.now() - timedelta(days=30)
            queryset = queryset.filter(created_at__gte=start_date)
        elif date_range == 'last_90_days':
            start_date = timezone.now() - timedelta(days=90)
            queryset = queryset.filter(created_at__gte=start_date)
        elif date_range == 'custom' and custom_start_date and custom_end_date:
            try:
                start_date = datetime.strptime(custom_start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                end_date = datetime.strptime(custom_end_date, '%Y-%m-%d').replace(tzinfo=timezone.utc) + timedelta(days=1)
                queryset = queryset.filter(created_at__gte=start_date, created_at__lt=end_date)
                logger.info(f"Custom date range: {start_date} to {end_date}")
            except ValueError as e:
                logger.error(f"Invalid date format: {e}")
        
        # Additional filters
        if filters.get('form_schema'):
            queryset = queryset.filter(form_schema_id=filters['form_schema'])
        
        if filters.get('status'):
            status_filter = filters['status']
            if status_filter == 'completed':
                queryset = queryset.filter(is_completed=True)
            elif status_filter == 'verified':
                queryset = queryset.filter(is_verified=True)
            elif status_filter == 'pending':
                queryset = queryset.filter(is_completed=False, is_verified=False)
        
        if filters.get('search'):
            search_term = filters['search']
            queryset = queryset.filter(
                Q(employee__first_name__icontains=search_term) |
                Q(employee__last_name__icontains=search_term) |
                Q(form_schema__name__icontains=search_term) |
                Q(form_data__icontains=search_term)
            )
        
        return queryset.select_related('employee', 'form_schema', 'organization')
    
    def export_to_excel(self, entries, options, file_name):
        """Export to Excel with clean, concise format showing only essential fields"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Entries"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Get all unique field names from schemas to create dynamic columns
        schema_fields = set()
        for entry in entries:
            if entry.form_schema and entry.form_schema.fields_definition:
                for field in entry.form_schema.fields_definition:
                    if isinstance(field, dict) and 'name' in field:
                        schema_fields.add(field['name'])
        
        # Create headers
        headers = ['Case ID', 'Employee', 'Status', 'Created Date']
        # Add schema fields to headers
        for field_name in sorted(schema_fields):
            headers.append(field_name.replace('_', ' ').title())
        headers.append('Files')
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = border
        
        # Write data
        for row, entry in enumerate(entries, 2):
            col = 1
            
            # Case ID
            ws.cell(row=row, column=col, value=str(entry.case_id or entry.entry_id))
            col += 1
            
            # Employee
            ws.cell(row=row, column=col, value=f"{entry.employee.first_name} {entry.employee.last_name}")
            col += 1
            
            # Status with color coding
            status_cell = ws.cell(row=row, column=col, value=self.get_status_text(entry))
            if 'Out of TAT' in self.get_status_text(entry):
                status_cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                status_cell.font = Font(color="FFFFFF", bold=True)
            col += 1
            
            # Created Date
            ws.cell(row=row, column=col, value=entry.created_at.strftime('%Y-%m-%d'))
            col += 1
            
            # Add schema field values
            form_data = entry.form_data or {}
            for field_name in sorted(schema_fields):
                value = form_data.get(field_name, '')
                ws.cell(row=row, column=col, value=str(value))
                col += 1
            
            # Files
            file_attachments = FileAttachment.objects.filter(form_entry=entry)
            form_field_files = FormFieldFile.objects.filter(form_entry=entry)
            
            file_list = []
            for attachment in file_attachments:
                file_list.append(attachment.original_filename)
            
            for field_file in form_field_files:
                file_list.append(f"{field_file.field_name}: {field_file.original_filename}")
            
            file_text = ", ".join(file_list) if file_list else "No files"
            ws.cell(row=row, column=col, value=file_text)
            
            # Apply borders to all cells in the row
            for c in range(1, col + 1):
                ws.cell(row=row, column=c).border = border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to response
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
        wb.save(response)
        return response
    
    def export_to_pdf(self, entries, options, file_name):
        """Export to PDF with auto-scaling for optimal readability"""
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{file_name}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=A4)
        story = []
        
        # Title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph("Form Entries Report", title_style))
        story.append(Spacer(1, 20))
        
        # Summary Statistics (simplified)
        total_entries = entries.count()
        completed_entries = entries.filter(is_completed=True).count()
        verified_entries = entries.filter(is_verified=True).count()
        out_of_tat_entries = sum(1 for entry in entries if entry.is_out_of_tat)
        
        summary_data = [
            ['Metric', 'Count', 'Percentage'],
            ['Total Entries', str(total_entries), '100%'],
            ['Completed', str(completed_entries), f"{(completed_entries/total_entries*100):.1f}%" if total_entries > 0 else "0%"],
            ['Verified', str(verified_entries), f"{(verified_entries/total_entries*100):.1f}%" if total_entries > 0 else "0%"],
            ['Out of TAT', str(out_of_tat_entries), f"{(out_of_tat_entries/total_entries*100):.1f}%" if total_entries > 0 else "0%"]
        ]
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 1*inch, 1*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 30))
        
        # Get all unique field names from schemas to create dynamic columns
        schema_fields = set()
        for entry in entries:
            if entry.form_schema and entry.form_schema.fields_definition:
                for field in entry.form_schema.fields_definition:
                    if isinstance(field, dict) and 'name' in field:
                        schema_fields.add(field['name'])
        
        # Create table headers
        headers = ['Case ID', 'Employee', 'Status', 'Created Date']
        for field_name in sorted(schema_fields):
            headers.append(field_name.replace('_', ' ').title())
        
        table_data = [headers]
        
        def extract_filename_from_url(url_string):
            """Extract filename from URL or return original string"""
            if isinstance(url_string, str) and url_string.startswith('http'):
                # Extract filename from URL
                try:
                    from urllib.parse import urlparse
                    parsed_url = urlparse(url_string)
                    path = parsed_url.path
                    filename = path.split('/')[-1]
                    if filename:
                        return filename, url_string  # Return both filename and original URL
                except:
                    pass
                return "File uploaded", url_string
            return url_string, None
        
        # Collect all data first to calculate optimal column widths
        all_rows_data = []
        max_content_lengths = [len(header) for header in headers]
        
        for entry in entries:
            form_data = entry.form_data or {}
            
            row_data = [
                str(entry.case_id or entry.entry_id),
                f"{entry.employee.first_name} {entry.employee.last_name}",
                self.get_status_text(entry),
                entry.created_at.strftime('%Y-%m-%d'),
            ]
            
            for field_name in sorted(schema_fields):
                value = form_data.get(field_name, '')
                original_url = None
                
                # Handle file fields - extract filename from URLs
                if isinstance(value, str) and value.startswith('http'):
                    value, original_url = extract_filename_from_url(value)
                
                # Truncate long values for readability
                if isinstance(value, str) and len(value) > 20:
                    value = value[:17] + "..."
                
                row_data.append(str(value))
            
            all_rows_data.append(row_data)
            
            # Update max content lengths for auto-scaling
            for i, content in enumerate(row_data):
                max_content_lengths[i] = max(max_content_lengths[i], len(str(content)))
        
        # Auto-scale column widths based on content
        num_columns = len(headers)
        available_width = 7.2 * inch  # A4 width minus margins
        min_column_width = 0.6 * inch
        max_column_width = 2.0 * inch
        
        # Calculate optimal column widths
        col_widths = []
        total_content_width = sum(max_content_lengths)
        
        for i, max_length in enumerate(max_content_lengths):
            # Calculate proportional width based on content length
            proportional_width = (max_length / total_content_width) * available_width
            
            # Apply min/max constraints
            optimal_width = max(min_column_width, min(max_column_width, proportional_width))
            col_widths.append(optimal_width)
        
        # Adjust if total width exceeds available space
        total_width = sum(col_widths)
        if total_width > available_width:
            # Scale down proportionally
            scale_factor = available_width / total_width
            col_widths = [width * scale_factor for width in col_widths]
        
        # Create table with auto-scaled data
        for row_data in all_rows_data:
            row = []
            for i, content in enumerate(row_data):
                # Create clickable link for file fields (check if it's a URL)
                if i >= 4:  # Schema fields start from index 4
                    field_name = list(sorted(schema_fields))[i - 4]
                    original_value = form_data.get(field_name, '')
                    
                    if isinstance(original_value, str) and original_value.startswith('http'):
                        link_style = ParagraphStyle(
                            'LinkStyle',
                            parent=styles['Normal'],
                            fontSize=6,
                            textColor=colors.blue,
                            underline=True
                        )
                        cell_content = Paragraph(f'<link href="{original_value}">{content}</link>', link_style)
                    else:
                        cell_content = content
                else:
                    cell_content = content
                
                row.append(cell_content)
            
            table_data.append(row)
        
        # Auto-scale font sizes based on number of columns
        if num_columns <= 6:
            header_font_size = 9
            data_font_size = 7
        elif num_columns <= 8:
            header_font_size = 8
            data_font_size = 6
        else:
            header_font_size = 7
            data_font_size = 5
        
        entries_table = Table(table_data, colWidths=col_widths)
        entries_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), header_font_size),
            ('FONTSIZE', (0, 1), (-1, -1), data_font_size),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('WORDWRAP', (0, 0), (-1, -1), True),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.beige, colors.white]),
        ]))
        story.append(entries_table)
        
        doc.build(story)
        return response
    
    def export_to_csv(self, entries, options, file_name):
        """Export to CSV with clean, concise format showing only essential fields"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'
        
        writer = csv.writer(response)
        
        # Get all unique field names from schemas to create dynamic columns
        schema_fields = set()
        for entry in entries:
            if entry.form_schema and entry.form_schema.fields_definition:
                for field in entry.form_schema.fields_definition:
                    if isinstance(field, dict) and 'name' in field:
                        schema_fields.add(field['name'])
        
        # Create headers
        headers = ['Case ID', 'Employee', 'Status', 'Created Date']
        # Add schema fields to headers
        for field_name in sorted(schema_fields):
            headers.append(field_name.replace('_', ' ').title())
        headers.append('Files')
        
        writer.writerow(headers)
        
        # Data
        for entry in entries:
            # Get file attachments for this entry
            file_attachments = FileAttachment.objects.filter(form_entry=entry)
            form_field_files = FormFieldFile.objects.filter(form_entry=entry)
            
            # Create concise file list
            file_list = []
            for attachment in file_attachments:
                file_list.append(attachment.original_filename)
            
            for field_file in form_field_files:
                file_list.append(f"{field_file.field_name}: {field_file.original_filename}")
            
            file_text = ", ".join(file_list) if file_list else "No files"
            
            # Get form data
            form_data = entry.form_data or {}
            
            # Create row data
            row = [
                str(entry.case_id or entry.entry_id),  # Case ID
                f"{entry.employee.first_name} {entry.employee.last_name}",  # Employee
                self.get_status_text(entry),  # Status
                entry.created_at.strftime('%Y-%m-%d'),  # Created Date
            ]
            
            # Add schema field values
            for field_name in sorted(schema_fields):
                value = form_data.get(field_name, '')
                row.append(str(value))
            
            row.append(file_text)  # Files
            writer.writerow(row)
        
        return response
    
    def get_status_text(self, entry):
        """Get status text for an entry"""
        if entry.is_completed:
            return 'Completed'
        elif entry.check_tat_status():
            return 'Out of TAT'
        else:
            return 'Pending'
    
    def create_summary_sheet(self, ws, entries):
        """Create summary sheet in Excel"""
        ws.title = "Summary"
        
        # Summary statistics
        total_entries = entries.count()
        completed_entries = entries.filter(is_completed=True).count()
        verified_entries = entries.filter(is_verified=True).count()
        out_of_tat_entries = entries.filter(
            is_completed=False,
            tat_start_time__lt=timezone.now() - timedelta(hours=24)
        ).count()
        
        summary_data = [
            ['Metric', 'Count'],
            ['Total Entries', total_entries],
            ['Completed Entries', completed_entries],
            ['Verified', verified_entries],
            ['Out of TAT', out_of_tat_entries],
            ['Completion Rate', f"{(completed_entries/total_entries*100):.1f}%" if total_entries > 0 else "0%"]
        ]
        
        for row, data in enumerate(summary_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    def create_analytics_sheet(self, ws, entries):
        """Create analytics sheet in Excel"""
        ws.title = "Analytics"
        
        # Analytics data
        analytics_data = [
            ['Analysis', 'Value'],
            ['Average TAT (hours)', ''],
            ['Fastest Completion (hours)', ''],
            ['Slowest Completion (hours)', ''],
            ['Most Active Employee', ''],
            ['Most Used Schema', '']
        ]
        
        # Calculate analytics
        completed_entries = entries.filter(is_completed=True)
        if completed_entries.exists():
            # Calculate using Python since tat_duration is a property
            tat_values = [entry.tat_duration for entry in completed_entries if entry.tat_duration]
            avg_tat = sum(tat_values) / len(tat_values) if tat_values else None
            fastest = min(tat_values) if tat_values else None
            slowest = max(tat_values) if tat_values else None
            
            analytics_data[1][1] = f"{avg_tat:.2f}" if avg_tat else "N/A"
            analytics_data[2][1] = f"{fastest:.2f}" if fastest else "N/A"
            analytics_data[3][1] = f"{slowest:.2f}" if slowest else "N/A"
        
        # Most active employee
        most_active_employee = entries.values('employee__first_name', 'employee__last_name').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        if most_active_employee:
            analytics_data[4][1] = f"{most_active_employee['employee__first_name']} {most_active_employee['employee__last_name']}"
        
        # Most used schema
        most_used_schema = entries.values('form_schema__name').annotate(
            count=Count('id')
        ).order_by('-count').first()
        
        if most_used_schema:
            analytics_data[5][1] = most_used_schema['form_schema__name']
        
        for row, data in enumerate(analytics_data, 1):
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                if row == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")

class EnhancedFormEntryExportView(APIView):
    """
    Enhanced export functionality for form entries with advanced filtering
    Supports Excel, PDF, and CSV with date range filtering and case ID tracking
    """
    permission_classes = [IsAuthenticated]
    
    def post(self, request):
        """Export form entries with advanced filtering"""
        user = request.user
        export_format = request.data.get('format', 'excel')
        filters = request.data.get('filters', {})
        options = request.data.get('options', {})
        
        logger.info(f"Enhanced export request from user {user.email}: format={export_format}, filters={filters}")
        
        # Get filtered entries
        if user.role == 'SUPER_ADMIN':
            organization = None
        else:
            organization = user.organization
        
        entries = self.get_filtered_entries(filters, organization)
        logger.info(f"Found {entries.count()} entries for export")
        
        # Generate filename with date range
        timestamp = timezone.now().strftime('%Y%m%d_%H%M%S')
        date_range = self.get_date_range_text(filters)
        file_name = f"form_entries_{date_range}_{timestamp}"
        
        if export_format == 'excel':
            return self.export_to_excel(entries, options, file_name)
        elif export_format == 'pdf':
            return self.export_to_pdf(entries, options, file_name)
        elif export_format == 'csv':
            return self.export_to_csv(entries, options, file_name)
        else:
            return Response({'error': 'Unsupported export format'}, status=status.HTTP_400_BAD_REQUEST)
    
    def get_date_range_text(self, filters):
        """Generate date range text for filename"""
        date_range = filters.get('date_range', 'all')
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        month = filters.get('month')
        year = filters.get('year')
        
        if date_range == 'today':
            return 'today'
        elif date_range == 'week':
            return 'last_7_days'
        elif date_range == 'month':
            return 'last_30_days'
        elif date_range == 'quarter':
            return 'last_90_days'
        elif date_range == 'year':
            return 'last_365_days'
        elif date_range == 'custom' and start_date and end_date:
            return f"{start_date}_to_{end_date}"
        elif start_date and end_date:
            return f"{start_date}_to_{end_date}"
        elif start_date:
            return f"from_{start_date}"
        elif end_date:
            return f"until_{end_date}"
        elif month and year:
            return f"{year}_month_{month}"
        elif year:
            return f"year_{year}"
        else:
            return 'all_entries'
    
    def get_filtered_entries(self, filters, organization):
        """Get filtered entries with comprehensive date range support"""
        queryset = FormEntry.objects.all()
        
        # Filter by organization
        if organization:
            queryset = queryset.filter(organization=organization)
        
        # Comprehensive date range filtering
        date_range = filters.get('date_range', 'all')
        start_date = filters.get('start_date')
        end_date = filters.get('end_date')
        month = filters.get('month')
        year = filters.get('year')
        
        # Handle different date range options
        if date_range == 'today':
            today_start = timezone.now().replace(hour=0, minute=0, second=0, microsecond=0)
            today_end = today_start + timedelta(days=1)
            queryset = queryset.filter(created_at__gte=today_start, created_at__lt=today_end)
        elif date_range == 'week':
            week_ago = timezone.now() - timedelta(days=7)
            queryset = queryset.filter(created_at__gte=week_ago)
        elif date_range == 'month':
            month_ago = timezone.now() - timedelta(days=30)
            queryset = queryset.filter(created_at__gte=month_ago)
        elif date_range == 'quarter':
            quarter_ago = timezone.now() - timedelta(days=90)
            queryset = queryset.filter(created_at__gte=quarter_ago)
        elif date_range == 'year':
            year_ago = timezone.now() - timedelta(days=365)
            queryset = queryset.filter(created_at__gte=year_ago)
        elif date_range == 'custom':
            # Handle custom date range with start_date and end_date
            if start_date:
                try:
                    start_datetime = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                    queryset = queryset.filter(created_at__gte=start_datetime)
                except ValueError as e:
                    logger.error(f"Invalid start_date format: {e}")
            
            if end_date:
                try:
                    end_datetime = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=timezone.utc) + timedelta(days=1)
                    queryset = queryset.filter(created_at__lt=end_datetime)
                except ValueError as e:
                    logger.error(f"Invalid end_date format: {e}")
        
        # Handle individual date filters (for advanced filters)
        if start_date and not date_range == 'custom':
            try:
                start_datetime = datetime.strptime(start_date, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                queryset = queryset.filter(created_at__gte=start_datetime)
            except ValueError as e:
                logger.error(f"Invalid start_date format: {e}")
        
        if end_date and not date_range == 'custom':
            try:
                end_datetime = datetime.strptime(end_date, '%Y-%m-%d').replace(tzinfo=timezone.utc) + timedelta(days=1)
                queryset = queryset.filter(created_at__lt=end_datetime)
            except ValueError as e:
                logger.error(f"Invalid end_date format: {e}")
        
        # Month and Year filtering
        if month is not None:
            queryset = queryset.filter(created_at__month=month)
        
        if year is not None:
            queryset = queryset.filter(created_at__year=year)
        
        # Advanced business filters
        if filters.get('bank_nbfc_name'):
            queryset = queryset.filter(form_data__bank_nbfc_name__icontains=filters['bank_nbfc_name'])
        
        if filters.get('location'):
            queryset = queryset.filter(form_data__location__icontains=filters['location'])
        
        if filters.get('product_type'):
            queryset = queryset.filter(form_data__product_type__icontains=filters['product_type'])
        
        if filters.get('case_status'):
            queryset = queryset.filter(form_data__case_status__icontains=filters['case_status'])
        
        # Personnel filters
        if filters.get('field_verifier_name'):
            queryset = queryset.filter(form_data__field_verifier_name__icontains=filters['field_verifier_name'])
        
        if filters.get('back_office_executive_name'):
            queryset = queryset.filter(form_data__back_office_executive_name__icontains=filters['back_office_executive_name'])
        
        # Status filtering
        if filters.get('status'):
            status_filter = filters['status']
            if status_filter == 'completed':
                queryset = queryset.filter(is_completed=True)
            elif status_filter == 'verified':
                queryset = queryset.filter(is_verified=True)
            elif status_filter == 'pending':
                queryset = queryset.filter(is_completed=False, is_verified=False)
        
        # Employee filtering
        if filters.get('employee_name'):
            queryset = queryset.filter(
                Q(employee__first_name__icontains=filters['employee_name']) |
                Q(employee__last_name__icontains=filters['employee_name'])
            )
        
        # Form schema filtering
        if filters.get('form_schema'):
            queryset = queryset.filter(form_schema_id=filters['form_schema'])
        
        # Search filtering
        if filters.get('search'):
            search_term = filters['search']
            queryset = queryset.filter(
                Q(employee__first_name__icontains=search_term) |
                Q(employee__last_name__icontains=search_term) |
                Q(employee__email__icontains=search_term) |
                Q(form_schema__name__icontains=search_term) |
                Q(verification_notes__icontains=search_term) |
                Q(case_id__icontains=search_term) |
                Q(form_data__icontains=search_term)
            )
        
        # TAT filtering (Out of TAT)
        if filters.get('is_out_of_tat') is not None:
            is_out_of_tat = filters['is_out_of_tat']
            if isinstance(is_out_of_tat, str):
                is_out_of_tat = is_out_of_tat.lower() == 'true'
            else:
                is_out_of_tat = bool(is_out_of_tat)
            
            # Filter for entries that are out of TAT using schema-specific limits
            out_of_tat_entries = []
            for entry in queryset:
                if entry.check_tat_status() == is_out_of_tat:
                    out_of_tat_entries.append(entry.id)
            queryset = queryset.filter(id__in=out_of_tat_entries)
        
        # Repeat case filtering
        if filters.get('is_repeat_case') is not None:
            is_repeat = filters['is_repeat_case']
            if isinstance(is_repeat, str):
                is_repeat = is_repeat.lower() == 'true'
            else:
                is_repeat = bool(is_repeat)
            queryset = queryset.filter(form_data__is_repeat_case=is_repeat)
        
        logger.info(f"Enhanced filtered entries count: {queryset.count()}")
        return queryset.select_related('employee', 'form_schema', 'organization')
    
    def export_to_excel(self, entries, options, file_name):
        """Export to Excel with enhanced formatting and case ID tracking"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Form Entries"
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Headers
        headers = [
            "Case ID", "Organization", "Employee", "Form Schema", "Status", 
            "Created Date", "Completed Date", "Verified Date", "TAT Status",
            "Bank/NBFC", "Location", "Product Type", "Case Status",
            "Field Verifier", "Back Office Executive", "Repeat Case",
            "Form Data", "Verification Notes", "File Attachments"
        ]
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write data
        for row, entry in enumerate(entries, 2):
            # Get form data
            form_data = entry.form_data or {}
            
            # Get file attachments
            attachments = entry.attachments.all()
            attachment_links = []
            for attachment in attachments:
                if hasattr(attachment, 'file') and attachment.file:
                    # Try to get S3 URL
                    try:
                        s3_url = S3FileManager.get_presigned_url(attachment.file.name)
                        attachment_links.append(f"{attachment.original_filename}: {s3_url}")
                    except:
                        attachment_links.append(f"{attachment.original_filename}: File not accessible")
            
            row_data = [
                entry.case_id or "N/A",
                entry.organization.display_name if entry.organization else "N/A",
                f"{entry.employee.first_name} {entry.employee.last_name}" if entry.employee else "N/A",
                entry.form_schema.name if entry.form_schema else "N/A",
                self.get_status_text(entry),
                entry.created_at.strftime('%Y-%m-%d %H:%M:%S') if entry.created_at else "N/A",
                entry.tat_completion_time.strftime('%Y-%m-%d %H:%M:%S') if entry.tat_completion_time else "N/A",
                entry.verified_at.strftime('%Y-%m-%d %H:%M:%S') if entry.verified_at else "N/A",
                "Out of TAT" if entry.is_out_of_tat else "Within TAT",
                form_data.get('bank_nbfc_name', 'N/A'),
                form_data.get('location', 'N/A'),
                form_data.get('product_type', 'N/A'),
                form_data.get('case_status', 'N/A'),
                form_data.get('field_verifier_name', 'N/A'),
                form_data.get('back_office_executive_name', 'N/A'),
                "Yes" if form_data.get('is_repeat_case') else "No",
                json.dumps(form_data, indent=2) if form_data else "N/A",
                entry.verification_notes or "N/A",
                "; ".join(attachment_links) if attachment_links else "No attachments"
            ]
            
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Create summary sheet
        self.create_summary_sheet(wb, entries)
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{file_name}.xlsx"'
        return response
    
    def export_to_pdf(self, entries, options, file_name):
        """Export to PDF with enhanced formatting"""
        doc = SimpleDocTemplate(
            BytesIO(),
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=72
        )
        
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        story.append(Paragraph(f"Form Entries Report - {file_name}", title_style))
        story.append(Spacer(1, 12))
        
        # Summary
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=20
        )
        story.append(Paragraph(f"Total Entries: {entries.count()}", summary_style))
        story.append(Paragraph(f"Generated: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}", summary_style))
        story.append(Spacer(1, 20))
        
        # Table data
        table_data = [['Case ID', 'Employee', 'Status', 'Created', 'Bank/NBFC', 'Location']]
        
        for entry in entries:
            form_data = entry.form_data or {}
            table_data.append([
                str(entry.case_id or "N/A"),
                f"{entry.employee.first_name} {entry.employee.last_name}" if entry.employee else "N/A",
                self.get_status_text(entry),
                entry.created_at.strftime('%Y-%m-%d') if entry.created_at else "N/A",
                form_data.get('bank_nbfc_name', 'N/A'),
                form_data.get('location', 'N/A')
            ])
        
        # Create table
        table = Table(table_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(table)
        
        # Build PDF
        doc.build(story)
        pdf_content = doc.getvalue()
        
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{file_name}.pdf"'
        return response
    
    def export_to_csv(self, entries, options, file_name):
        """Export to CSV with enhanced data"""
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{file_name}.csv"'
        
        writer = csv.writer(response)
        
        # Headers
        headers = [
            "Case ID", "Organization", "Employee", "Form Schema", "Status", 
            "Created Date", "Completed Date", "Verified Date", "TAT Status",
            "Bank/NBFC", "Location", "Product Type", "Case Status",
            "Field Verifier", "Back Office Executive", "Repeat Case",
            "Verification Notes"
        ]
        writer.writerow(headers)
        
        # Data
        for entry in entries:
            form_data = entry.form_data or {}
            row = [
                entry.case_id or "N/A",
                entry.organization.display_name if entry.organization else "N/A",
                f"{entry.employee.first_name} {entry.employee.last_name}" if entry.employee else "N/A",
                entry.form_schema.name if entry.form_schema else "N/A",
                self.get_status_text(entry),
                entry.created_at.strftime('%Y-%m-%d %H:%M:%S') if entry.created_at else "N/A",
                entry.tat_completion_time.strftime('%Y-%m-%d %H:%M:%S') if entry.tat_completion_time else "N/A",
                entry.verified_at.strftime('%Y-%m-%d %H:%M:%S') if entry.verified_at else "N/A",
                "Out of TAT" if entry.is_out_of_tat else "Within TAT",
                form_data.get('bank_nbfc_name', 'N/A'),
                form_data.get('location', 'N/A'),
                form_data.get('product_type', 'N/A'),
                form_data.get('case_status', 'N/A'),
                form_data.get('field_verifier_name', 'N/A'),
                form_data.get('back_office_executive_name', 'N/A'),
                "Yes" if form_data.get('is_repeat_case') else "No",
                entry.verification_notes or "N/A"
            ]
            writer.writerow(row)
        
        return response
    
    def get_status_text(self, entry):
        """Get status text for entry"""
        if entry.is_verified:
            return "Verified"
        elif entry.is_completed:
            return "Completed"
        else:
            return "Pending"
    
    def create_summary_sheet(self, wb, entries):
        """Create summary sheet with analytics"""
        ws = wb.create_sheet("Summary")
        
        # Summary statistics
        total_entries = entries.count()
        completed_entries = entries.filter(is_completed=True).count()
        verified_entries = entries.filter(is_verified=True).count()
        pending_entries = entries.filter(is_completed=False, is_verified=False).count()
        
        # Write summary
        ws['A1'] = "Form Entries Summary"
        ws['A1'].font = Font(bold=True, size=14)
        
        ws['A3'] = "Total Entries"
        ws['B3'] = total_entries
        
        ws['A4'] = "Completed"
        ws['B4'] = completed_entries
        
        ws['A5'] = "Verified"
        ws['B5'] = verified_entries
        
        ws['A6'] = "Pending"
        ws['B6'] = pending_entries
        
        # Calculate percentages
        if total_entries > 0:
            ws['A8'] = "Completion Rate"
            ws['B8'] = f"{(completed_entries/total_entries)*100:.1f}%"
            
            ws['A9'] = "Verification Rate"
            ws['B9'] = f"{(verified_entries/total_entries)*100:.1f}%"
